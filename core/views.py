from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth import login, logout
from django.contrib import messages
from django.contrib.auth.models import User

from django.contrib.auth.decorators import login_required



from django.utils import timezone   # ✅ مهم عشان نحسب المهام المتأخرة
from .models import Project, Task, Profile, Submission, Project, ActivityLog, Skill
from .forms import CustomUserCreationForm, SubmissionForm, ProjectForm


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from .ai import chat_with_stad_ai, suggest_assignee_for_task, suggest_assignee_for_form_input
import json

# ✅ للتصدير إلى Excel

from django.db import models
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import datetime

# بنعرّفها بعد شوي أو عدّل الاسم حسب ai.py عندك




from .forms import CustomUserCreationForm





@login_required
def test_ai_view(request):
    """
    View بسيطة بس نجرب من خلالها الربط مع OpenAI.
    ترجع رد AI كـ JSON.
    """
    user_message = "السلام عليكم"
    
    try:
        from .ai import chat_with_stad_ai  # تأكد الاسم نفس اللي في ai.py

        ai_reply = chat_with_stad_ai(user_message)
    except Exception as e:
        ai_reply = f"AI ERROR: {e}"

    return JsonResponse(
    {
        "input": user_message,
        "reply": ai_reply,
    },
    json_dumps_params={"ensure_ascii": False},  # ✅ عشان يطلع العربي صح
)

 
@csrf_exempt   # نسهّل الاختبار الآن، بعدين نضبط CSRF مع الفرونت
@login_required
def smart_assign_api(request):
    """
    API لزر 'إسناد ذكي' في شاشة إنشاء المهمة.
    - ما ينشئ مهمة
    - ما يغيّر الداتابيس
    - بس يرجّع أفضل موظف مقترح + السبب + السكور
    """
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    title = data.get("title", "").strip()
    description = data.get("description", "").strip()
    due_date = data.get("due_date")  # ممكن تكون None أو "2025-12-31"
    required_skill_ids = data.get("required_skills", [])  # list of ints

    if not title and not description:
        return JsonResponse({"error": "يجب إدخال عنوان أو وصف للمهمة"}, status=400)

    try:
        result = suggest_assignee_for_form_input(
            title=title,
            description=description,
            due_date=due_date,
            required_skill_ids=required_skill_ids,
            priority=data.get("priority", "medium"),  
            impact=data.get("impact", "normal"),
        )
    except Exception as e:
        return JsonResponse({"error": f"AI error: {e}"}, status=500)

    assigned_id = result.get("assigned_user_id")
    reason = result.get("reason", "")
    scores = result.get("scores", [])

    suggested_user = None
    if assigned_id:
        from django.contrib.auth.models import User
        from .models import Profile

        user = (
            User.objects
            .filter(id=assigned_id)
            .select_related("profile")
            .first()
        )
        if user:
            profile = getattr(user, "profile", None)
            suggested_user = {
                "id": user.id,
                "name": user.get_full_name() or user.username,
                "job_role": profile.job_role.name if profile and profile.job_role else None,
                "overall_rating": profile.overall_rating if profile else 0,
                "rating_count": profile.rating_count if profile else 0,
                "current_workload": profile.calculate_workload() if profile else 0,
            }

    return JsonResponse(
        {
            "ok": True,
            "suggested_user": suggested_user,
            "reason": reason,
            "scores": scores,
        },
        json_dumps_params={"ensure_ascii": False},
    )


# Create your views here.
def home(request):
    return render(request, 'home.html')

def login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            # بعدين نودّيه للداشبورد، الآن نخليه يرجع للهوم
            return redirect('dashboard')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {"form": form})

def register_view(request):
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST)  # ✅ تغيير هنا
        if form.is_valid():
            user = form.save()
            login(request, user)  # يسجّله دخول مباشرة بعد التسجيل
            return redirect('dashboard')
    else:
        form = CustomUserCreationForm()  # ✅ وهنا برضو
    return render(request, 'register.html', {"form": form})

def logout_view(request):
    logout(request)
    return redirect('home')


@login_required
def dashboard(request):
    user = request.user

    profile = Profile.objects.filter(user=user).first()
    selected_employee = None
    role = profile.role if profile else "employee"
    today = timezone.localdate()
    
        # ========== تحديث المهام المتأخرة ==========
    for task in Task.objects.filter(due_date__isnull=False):
        if task.status not in ["completed", "late","under_review"]:
            if task.due_date < today:
                task.status = "late"
                task.save()
    late_qs = Task.objects.filter(
        due_date__isnull=False,
        due_date__lt=today,      # المهام التي انتهى وقتها فعلاً
    ).exclude(status__in=["completed", "late", "under_review"])

    if late_qs.exists():
        late_qs.update(status="late")

    # ===========================================


    

    projects = Project.objects.none()
    tasks = Task.objects.none()
    employees = []
    submit_tasks = None
    status_filter = ""
    project_filter = ""
    selected_employee = None

    total_tasks = 0
    completed_tasks = 0
    in_progress_tasks = 0
    overdue_tasks = 0
    active_projects = 0
    workload = 0

    todo_tasks = 0
    active_tasks = 0
    workload_capacity = 10
    under_review_tasks = 0
    
    if role == "manager":
        # المدير يشوف كل المشاريع
        projects = Project.objects.all()
        

        # 🔹 Query أساسي لكل المهام (يُستخدم للإحصائيات والكروت)
        base_tasks = Task.objects.all().select_related("assignee", "project")
        

        # 🔹 هذا اللي بنعرضه في جدول "المهام" وبيتغيّر حسب الفلاتر
        tasks = base_tasks

        # قراءة فلاتر GET
        # قراءة فلاتر GET
        status_filter = request.GET.get("status", "")
        selected_employee = request.GET.get("employee", "")

        # فلتر الحالة ← يؤثر فقط على جدول المهام
        if status_filter in ["todo", "in_progress", "under_review", "completed", "blocked", "late"]:
            tasks = tasks.filter(status=status_filter)
        
        # فلتر الموظف ← يؤثر فقط على جدول المهام
        if selected_employee:
            tasks = tasks.filter(assignee__id=selected_employee)
        
        # قائمة الموظفين للـ Dropdown في نموذج إضافة مهمة
        employees = User.objects.filter(
            profile__role="employee"
        ).select_related("profile").order_by("username")


        # 🔸 الإحصائيات تعتمد على كل المهام (base_tasks) وليس tasks المفلترة
        total_tasks = base_tasks.count()
        completed_tasks = base_tasks.filter(status="completed").count()
        in_progress_tasks = base_tasks.filter(status="in_progress").count()
        under_review_tasks = base_tasks.filter(status="under_review").count()


        # المهام المتأخرة من كل المهام
        overdue_tasks = base_tasks.filter(
            due_date__isnull=False,
            due_date__lt=today,
        ).exclude(status="completed").count()

        # عدد كل المشاريع
        active_projects = projects.count()
        
        # نسبة إنجاز المشاريع تحسب من كل المهام (مو المفلترة)
        for project in projects:
            proj_tasks = base_tasks.filter(project=project)
            project.task_count = proj_tasks.count()

            completed_in_proj = proj_tasks.filter(status="completed").count()

            if project.task_count > 0:
                project.completion = int(completed_in_proj * 100 / project.task_count)
            else:
                project.completion = 0
        
    else:
    # ================= موظف =================

    # 1) نقرأ الفلاتر من الـ GET
        status_filter = request.GET.get("status", "")
        project_filter = request.GET.get("project", "")

    # 2) الـ Query الأساسي لكل مهام الموظف (نستخدمه للإحصائيات)
        base_qs = Task.objects.filter(assignee=user).select_related("project")

        late_tasks_qs = base_qs.filter(
            due_date__isnull=False,
            due_date__lt=today,
        ).exclude(status__in=["completed", "under_review"])

    # ⬅️ هذه المهام تستخدم فقط في قائمة "تسليم المهام"
        submit_tasks = base_qs.filter(status="in_progress")

    # 3) نطبّق الفلاتر على نسخة ثانية لجدول "مهامي"
        employee_tasks = base_qs

        if status_filter:
            if status_filter == "late":
                employee_tasks = employee_tasks.filter(status="late")
            else:
                employee_tasks = employee_tasks.filter(status=status_filter)



        if project_filter:
            employee_tasks = employee_tasks.filter(project_id=project_filter)

    # 4) الإحصائيات تبقى من الـ base_qs (كل المهام بدون فلترة)
        total_tasks = base_qs.count()
        completed_tasks = base_qs.filter(status="completed").count()
        in_progress_tasks = base_qs.filter(status="in_progress").count()
        under_review_tasks = base_qs.filter(status="under_review").count()

        todo_tasks = base_qs.filter(status="todo").count()
        active_tasks = todo_tasks + in_progress_tasks
        workload_capacity = 10
        workload = profile.current_workload or 0

        # ✅ إضافة بيانات التقييم للموظف
        overall_rating = profile.overall_rating if profile else 0.0
        rating_count = profile.rating_count if profile else 0
    # 5) اللي يروح للجدول هو employee_tasks بعد الفلترة
        tasks = employee_tasks

    # 6) المشاريع الخاصة بالموظف (مميّزة عشان نستخدمها في فلتر المشروع + مشاريعي)
        projects = Project.objects.filter(tasks__assignee=user).distinct()

    # 7) نحسب لكل مشروع عدد مهام الموظف ونسبة الإنجاز (من الـ base_qs)
        for project in projects:
            proj_tasks = base_qs.filter(project=project)
            project.emp_task_count = proj_tasks.count()
            completed = proj_tasks.filter(status="completed").count()

            if project.emp_task_count > 0:
                project.emp_completion = int(completed * 100 / project.emp_task_count)
            else:
                project.emp_completion = 0

        overdue_tasks = late_tasks_qs.count()
        active_projects = projects.count()
        if role != "manager":
            employees = []

    workload_employees = []
    all_profiles = Profile.objects.select_related("user", "job_role")

    for p in all_profiles:

        if p.role == "manager":
            continue

        workload_employees.append({
        "id": p.user.id,
        "name": p.user.first_name or p.user.username,
        "role": p.job_role.name if p.job_role else "—",
        "workload": p.current_workload,
        "level": (
            "low" if p.current_workload <= 3 else
            "medium" if p.current_workload <= 6 else
            "high"
        ),
    })
    

    activities = ActivityLog.objects.order_by("-timestamp")[:5]

     # 👇 أضف هذا السطر هنا
    skills = Skill.objects.all().order_by("name")
    
# euuewfewjiwoijfioejoifjiwoofeijoiowejifjeiowfiojeoiwjfoijweoijfoijewoifjoiewoifioeoifewoijfoijewoifjoiewjfoiejwoifoiewoifewfoieoifoiewfoijoiefoiwejfoijefeoifjwoifjoijeoifjoiejoifjwoiejfoiwjo
#    for p in all_profiles:
 #       employees.append({
  #          "name": p.user.first_name or p.user.username,
   #         "role": p.job_role.name if p.job_role else "—",
    #        "workload": p.current_workload,
     #       "level": (
      #          "low" if p.current_workload <= 3 else
       #         "medium" if p.current_workload <= 6 else
        #        "high"
         #   ),
        #})


    # ======================
    # آخر النشاطات
    # ======================
    

        



    




    context = {
        "role": role,
        "projects": projects,
        "tasks": tasks,
        "employees": employees,
        "total_tasks": total_tasks,
        "completed_tasks": completed_tasks,
        "in_progress_tasks": in_progress_tasks,
        "overdue_tasks": overdue_tasks,        # ✅ مهم للكرت
        "active_projects": active_projects,
        "submit_tasks": submit_tasks if role != "manager" else None,
        "selected_status": status_filter if role != "manager" else "",
        "selected_project": project_filter if role != "manager" else "",
        "activities": activities,
        "selected_status_manager": status_filter or "",
        "workload_employees": workload_employees,
        "selected_employee": selected_employee,
        "todo_tasks": todo_tasks,
        "active_tasks": active_tasks,
        "workload_capacity": workload_capacity,
        "skills": skills,
        "workload": workload,
        "under_review_tasks": under_review_tasks,

        "overall_rating": locals().get("overall_rating", 0.0),
        "rating_count": locals().get("rating_count", 0),


    }
    return render(request, "dashboard.html", context)

@login_required
def edit_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    # ✅ تحقق أن المستخدم مدير
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بتعديل المشاريع.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تعديل المشروع بنجاح ✅')
        else:
            messages.error(request, 'حدث خطأ في البيانات.')
        
        return redirect('dashboard')

    return redirect('dashboard')
@login_required
def delete_project(request, project_id):
    # ✅ نجيب المشروع
    project = get_object_or_404(Project, id=project_id)

    # ✅ نتأكد إن اللي يحذف = مدير
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بحذف المشاريع.")
        return redirect('dashboard')

    # ما نحذف إلا لو جاي الطلب من الفورم (POST)
    if request.method == "POST":
        project_name = project.name

        # لو الـ FK في Task = CASCADE يكفي هذا السطر:
        project.delete()

        # لو مو متأكد من CASCADE وتبي تضمن 100٪:
        #Task.objects.filter(project=project).delete()
        #project.delete()

        messages.success(
            request,
            f"تم حذف المشروع ({project_name}) وجميع المهام المرتبطة به بنجاح 🗑️"
        )
        return redirect('dashboard')

    # لو أحد فتح الرابط GET نرجعه للداشبورد
    return redirect('dashboard')

@login_required
def submit_task(request):
    if request.method == "POST":
        task_id = request.POST.get("task_id")
        notes = request.POST.get("notes", "")
        attachment = request.FILES.get("attachment")

        if not task_id:
            messages.error(request, "يجب اختيار مهمة أولاً.")
            return redirect("dashboard")

        # نتأكد أن المهمة فعلاً تابعة لهذا الموظف
        task = get_object_or_404(Task, id=task_id, assignee=request.user)

        Submission.objects.create(
            task=task,
            employee=request.user,
            notes=notes,
            attachment=attachment,
            status='under_review',
        )

        # لو حبيت نعتبر المهمة مكتملة بعد التسليم
        task.status = "under_review"
        task.save()

        messages.success(request, "تم تسليم المهمة بنجاح.")
        return redirect("dashboard")

    # أي GET على هذا الرابط نرجعه للداشبورد
    return redirect("dashboard")

# ===================================
# صفحة التسليمات (للمدير فقط)
# ===================================

@login_required
def submissions_view(request):
    """
    عرض جميع التسليمات (للمدير فقط)
    """
    # التحقق من أن المستخدم مدير
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بالوصول لهذه الصفحة.")
        return redirect('dashboard')

    # جلب جميع التسليمات
    submissions = Submission.objects.all().select_related(
        'task',
        'task__project',
        'employee'
    ).order_by('-submitted_at')

    # الفلاتر
    status_filter = request.GET.get('status', '')
    employee_filter = request.GET.get('employee', '')
    project_filter = request.GET.get('project', '')

    if status_filter:
        submissions = submissions.filter(status=status_filter)
    
    if employee_filter:
        submissions = submissions.filter(employee_id=employee_filter)
    
    if project_filter:
        submissions = submissions.filter(task__project_id=project_filter)

    # الإحصائيات
    all_submissions = Submission.objects.all()
    total_submissions = all_submissions.count()
    pending_submissions = all_submissions.filter(status='under_review').count()
    approved_submissions = all_submissions.filter(status='approved').count()
    rejected_submissions = all_submissions.filter(status='rejected').count()

    # قوائم للفلاتر
    employees = User.objects.filter(profile__role='employee').order_by('username')
    projects = Project.objects.all().order_by('name')

    context = {
        'submissions': submissions,
        'total_submissions': total_submissions,
        'pending_submissions': pending_submissions,
        'approved_submissions': approved_submissions,
        'rejected_submissions': rejected_submissions,
        'employees': employees,
        'projects': projects,
        'selected_status': status_filter,
        'selected_employee': employee_filter,
        'selected_project': project_filter,
    }

    return render(request, 'submissions.html', context)

@login_required
def approve_submission(request, submission_id):
    """
    الموافقة على تسليم مهمة
    ✅ الحل الذهبي: فقط التسليمات الموافق عليها تأثر على التقييم العام
    """
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بهذا الإجراء.")
        return redirect('dashboard')

    if request.method == "POST":
        submission = get_object_or_404(Submission, id=submission_id)
        
        # الحصول على التقييم والملاحظات
        rating = request.POST.get('rating', 5)
        comment = request.POST.get('comment', '')
        
        # ✅ تحديث التسليم
        submission.status = 'approved'
        submission.reviewed_at = timezone.now()
        submission.reviewed_by = request.user
        submission.rating = int(rating)  # حفظ التقييم في Submission
        submission.manager_comment = comment  # حفظ الملاحظات في Submission
        submission.save()

        # تحديث حالة المهمة إلى مكتملة
        task = submission.task
        task.status = 'completed'
        task.save()

        # ✅ فقط التسليمات الموافق عليها تسجّل في TaskRating
        # هذا يأثر على Profile.overall_rating
        from .models import TaskRating
        TaskRating.objects.create(
            task=task,
            employee=submission.employee,
            manager=request.user,
            rating=int(rating),
            comment=comment,
        )

        # تسجيل النشاط
        ActivityLog.objects.create(
            user=request.user,
            action_type="submission_approved",
            message=f"تمت الموافقة على تسليم المهمة ({task.title}) من {submission.employee.username} | التقييم: {rating}★",
        )

        messages.success(
            request,
            f"تمت الموافقة على التسليم بنجاح ✓ | التقييم: {rating} ⭐"
        )

    return redirect('submissions')


@login_required
def reject_submission(request, submission_id):
    """
    رفض تسليم مهمة
    ❌ الحل الذهبي: التسليمات المرفوضة لا تأثر على التقييم العام
    """
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بهذا الإجراء.")
        return redirect('dashboard')

    if request.method == "POST":
        submission = get_object_or_404(Submission, id=submission_id)
        
        # الحصول على ملاحظات الرفض
        comment = request.POST.get('comment', '')
        
        # ✅ تحديث التسليم
        submission.status = 'rejected'
        submission.reviewed_at = timezone.now()
        submission.reviewed_by = request.user
        submission.rating = 1  # تقييم منخفض للسجل فقط
        submission.manager_comment = comment
        submission.save()

        # ❌ لا نسجل في TaskRating (لا يأثر على التقييم العام)
        # فقط نحفظ في Submission للسجل

        # إعادة المهمة لحالة "قيد التنفيذ"
        task = submission.task
        task.status = 'in_progress'
        
        # ✅ تحديث وصف المهمة بملاحظات الرفض
        if comment:
            old_description = task.description if task.description else ""
            fnow = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            
            task.description = f"""🔴 ملاحظات المدير (مرفوض) - {now}
👤 المدير: {request.user.get_full_name() or request.user.username}
💬 الملاحظة: {comment}

{'─' * 50}

{old_description}"""
        
        task.save()

        # تسجيل النشاط
        ActivityLog.objects.create(
            user=request.user,
            action_type="submission_rejected",
            message=f"تم رفض تسليم المهمة ({task.title}) من {submission.employee.username}",
        )

        messages.warning(
            request,
            f"تم رفض التسليم. المهمة أُعيدت لحالة 'قيد التنفيذ' مع ملاحظاتك ✗"
        )

    return redirect('submissions')


@login_required
def create_project(request):
    # يسمح فقط للمدير
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بإضافة المشاريع.")
        return redirect('dashboard')

    if request.method == "POST":
        name        = request.POST.get("name")
        description = request.POST.get("description", "")
        start_date_str = request.POST.get("start_date")
        end_date_str   = request.POST.get("end_date")
        manager_id  = request.POST.get("manager") or request.user.id

        # نحول الـ ID إلى User
        manager = User.objects.get(id=manager_id)

        # نحول التواريخ من string إلى date (لو موجودة)
        start_date = None
        end_date   = None

        try:
            if start_date_str:
                start_date = datetime.date.fromisoformat(start_date_str)
            if end_date_str:
                end_date = datetime.date.fromisoformat(end_date_str)
        except ValueError:
            # لو تاريخ غلط نتجاهله ونكمل بدون ما نكسر الصفحة
            messages.warning(request, "صيغة التاريخ غير صحيحة، تم تجاهل التواريخ.")
        
        # إنشاء المشروع
        Project.objects.create(
            name=name,
            description=description,
            start_date=start_date,
            end_date=end_date,
            manager=manager,
        )

        messages.success(request, f"تم إضافة المشروع ({name}) بنجاح ✅")
        return redirect('dashboard')   # ✅ يرجّعك للداشبورد

    return redirect('dashboard')


@login_required
def create_task(request):
    # يسمح فقط للمدير
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بإضافة المهام.")
        return redirect('dashboard')

    if request.method == "POST":
        # جلب البيانات الأساسية
        title = request.POST.get("title")
        description = request.POST.get("description", "")
        project_id = request.POST.get("project")
        assignee_id = request.POST.get("assignee")   # 👈 هنا بيجينا الموظف المقترح أو اليدوي
        due_date_str = request.POST.get("due_date")
        status = request.POST.get("status", "todo")
        priority = request.POST.get("priority", "medium")
        impact = request.POST.get("impact", "normal")

        # (اختياري) لو حابين تخلون نوع الإسناد موجود في الفورم
        assignment_type = request.POST.get("assignment_type", "manual")
        # ⚠️ ملاحظة: ما راح نسوي أي منطق خاص لو كان "auto"
        # لأن الـ AI صار يشتغل في API منفصلة ويرجع فقط اقتراح

        # ✅ جلب المهارات المطلوبة
        required_skills_ids = request.POST.getlist("required_skills")

        # المشروع (اختياري)
        project = None
        if project_id:
            project = Project.objects.filter(id=project_id).first()

        # الموظف (اختياري)
        assignee = None
        if assignee_id:
            assignee = User.objects.filter(id=assignee_id).first()

        # التاريخ (اختياري)
        due_date = None
        if due_date_str:
            try:
                due_date = datetime.date.fromisoformat(due_date_str)
            except ValueError:
                pass

        # ✅ إنشاء المهمة (بدون تدخل AI نهائياً)
        task = Task.objects.create(
            title=title,
            description=description,
            project=project,
            assignee=assignee,
            due_date=due_date,
            status=status,
            impact=impact,        # 🔥 إضافة impact
            priority=priority,    # 🔥 إضافة priority الجديد
        )

        # ✅ إضافة المهارات المطلوبة
        if required_skills_ids:
            task.required_skills.set(required_skills_ids)

        # ✅ تسجيل النشاط
        ActivityLog.objects.create(
            user=request.user,
            action_type="task_created",
            message=f"تم إنشاء المهمة ({title})",
        )

        messages.success(request, f"تمت إضافة المهمة ({title}) بنجاح ✅")
        return redirect("dashboard")

    return redirect("dashboard")



@login_required
def start_task(request, task_id):
    
    task = get_object_or_404(Task, id=task_id, assignee=request.user)

    if task.status == "todo":
        task.status = "in_progress"
        task.save()
    return redirect("dashboard")

@login_required
def edit_task(request, task_id):
    # نجيب المهمة
    task = get_object_or_404(Task, id=task_id)

    # نتأكد أن اللي يعدل = مدير
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بتعديل المهام.")
        return redirect("dashboard")

    if request.method == "POST":
        # عنوان ووصف
        task.title = request.POST.get("title", task.title)
        task.description = request.POST.get("description", "")

        # المشروع
        project_id = request.POST.get("project")
        if project_id:
            task.project = get_object_or_404(Project, id=project_id)
        else:
            task.project = None

        # الموظف المكلّف
        assignee_id = request.POST.get("assignee")
        if assignee_id:
            task.assignee = get_object_or_404(User, id=assignee_id)
        else:
            task.assignee = None

        # الحالة
        status = request.POST.get("status")
        if status in ["todo", "in_progress","under_review", "completed", "blocked"]:
            task.status = status

        # تاريخ النهاية
        due_date_str = request.POST.get("due_date")
        if due_date_str:
            try:
                task.due_date = datetime.date.fromisoformat(due_date_str)
            except ValueError:
                # لو التاريخ غلط نتجاهله
                pass
        else:
            task.due_date = None

        task.save()

        # تسجيل نشاط (اختياري)
        ActivityLog.objects.create(
            user=request.user,
            action_type="task_updated",
            message=f"تم تعديل المهمة ({task.title})",
        )

        messages.success(request, "تم تعديل المهمة بنجاح ✅")
        return redirect("dashboard")

    # أي GET يرجعه للداشبورد
    return redirect("dashboard")


@login_required
def delete_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)

    # نتأكد أن اللي يحذف = مدير
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بحذف المهام.")
        return redirect("dashboard")

    if request.method == "POST":
        title = task.title
        task.delete()

        ActivityLog.objects.create(
            user=request.user,
            action_type="task_deleted",
            message=f"تم حذف المهمة ({title})",
        )

        messages.success(request, f"تم حذف المهمة ({title}) بنجاح 🗑️")
        return redirect("dashboard")

    return redirect("dashboard")



# ===================================
# تصدير الموظفين إلى Excel
# ===================================

@login_required
def export_employees_excel(request):
    """تصدير بيانات الموظفين - قسم IT"""
    # التحقق من الصلاحيات
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك بالوصول.")
        return redirect('dashboard')
    
    # إنشاء Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "موظفي IT"
    
    # ========== التنسيقات ==========
    title_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # ========== العنوان الرئيسي ==========
    ws.merge_cells('A1:K1')
    ws['A1'] = f'📊 تقرير موظفي قسم IT - {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 30
    
    # ========== العناوين ==========
    headers = [
        'ID',
        'اسم المستخدم',
        'الاسم الأول',
        'الاسم الأخير',
        'البريد الإلكتروني',
        'القسم',
        'المسمى الوظيفي',
        'المهارات',
        'التقييم العام',
        'عدد المهام المكتملة',
        'تاريخ الانضمام'
    ]
    
    ws.append([])  # سطر فارغ
    ws.append(headers)
    
    # تنسيق الرأس
    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    ws.row_dimensions[3].height = 25
    
    # ========== البيانات ==========
    employees = Profile.objects.filter(role='employee').select_related('user', 'job_role').prefetch_related('user__skill_set__skill')
    
    for emp in employees:
        # حساب الإحصائيات
        completed_tasks = Task.objects.filter(
            assignee=emp.user, 
            status='completed'
        ).count()
        
        # ✅ جلب المسمى الوظيفي من job_role وتحويله لـ string
        job_role_obj = getattr(emp, 'job_role', None)
        if job_role_obj:
            job_role = str(job_role_obj)
        else:
            job_role = 'موظف'
        
        # ✅ جلب المهارات باستخدام related_name
        skills = 'لا يوجد'
        try:
            # استخدام الـ related_name مباشرة
            employee_skills = emp.user.skill_set.select_related('skill').all()
            
            if employee_skills.exists():
                skills_list = []
                for es in employee_skills:
                    skill_name = str(es.skill.name)
                    skill_level = es.level
                    # عرض المهارة مع المستوى
                    skills_list.append(f"{skill_name} ({skill_level}/5)")
                
                skills = ', '.join(skills_list)
        except Exception as e:
            skills = f'خطأ: {str(e)}'
        
        # ✅ باقي البيانات
        department = 'IT'  # القسم دايماً IT
        
        # البيانات (بدون رقم الهاتف)
        row_data = [
            emp.id,
            str(emp.user.username),
            str(emp.user.first_name) if emp.user.first_name else 'غير محدد',
            str(emp.user.last_name) if emp.user.last_name else 'غير محدد',
            str(emp.user.email) if emp.user.email else 'غير محدد',
            department,
            job_role,
            skills,  # المهارات مع المستويات
            f"{emp.overall_rating:.2f} ⭐",
            completed_tasks,
            emp.user.date_joined.strftime('%Y-%m-%d')
        ]
        
        ws.append(row_data)
        
        # تنسيق السطر
        current_row = ws.max_row
        for cell in ws[current_row]:
            cell.alignment = data_alignment
            cell.border = border
            
            # تلوين حسب التقييم
            if cell.column == 9:  # عمود التقييم
                rating = emp.overall_rating
                if rating >= 4.5:
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif rating >= 3.5:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                elif rating > 0:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # ========== الإحصائيات في الأسفل (مع مسافة كبيرة) ==========
    
    # نضيف 10 أسطر فارغة عشان نبعد تماماً
    for _ in range(3):
        ws.append([])
    
    # صف العنوان
    stats_title_row = ws.max_row + 1
    ws[f'A{stats_title_row}'] = '📈 الإحصائيات العامة - قسم IT'
    ws[f'A{stats_title_row}'].font = Font(size=14, bold=True, color="FFFFFF")
    ws[f'A{stats_title_row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    ws[f'A{stats_title_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[stats_title_row].height = 30
    
    # دمج الخلايا بعد كتابة النص (على كل الأعمدة)
    ws.merge_cells(f'A{stats_title_row}:K{stats_title_row}')
    
    # سطر فارغ
    ws.append([])
    
    # الإحصائيات
    ws.append([
        'إجمالي الموظفين:',
        employees.count(),
        '',
        'متوسط التقييم العام:',
        f"{employees.aggregate(avg_rating=models.Avg('overall_rating'))['avg_rating'] or 0:.2f} ⭐",
        '',
        '',
        '',
        '',
        '',
        ''
    ])
    
    total_completed = sum(
        Task.objects.filter(assignee=emp.user, status='completed').count() 
        for emp in employees
    )
    
    ws.append([
        'إجمالي المهام المكتملة:',
        total_completed,
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        ''
    ])
    
    # ========== تعديل عرض الأعمدة ==========
    column_widths = {
        'A': 8,   # ID
        'B': 15,  # اسم المستخدم
        'C': 15,  # الاسم الأول
        'D': 15,  # الاسم الأخير
        'E': 25,  # البريد
        'F': 10,  # القسم
        'G': 20,  # المسمى الوظيفي
        'H': 50,  # المهارات (أوسع عشان تظهر المهارات مع المستويات)
        'I': 15,  # التقييم
        'J': 20,  # المهام المكتملة
        'K': 18,  # تاريخ الانضمام
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # ========== حفظ وإرجاع الملف ==========
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'IT_employees_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

@login_required
def export_tasks_excel(request):
    """تصدير المهام إلى Excel"""
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك.")
        return redirect('dashboard')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "المهام"
    
    # التنسيقات
    title_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # العنوان الرئيسي
    ws.merge_cells('A1:K1')
    ws['A1'] = f'📋 تقرير المهام - قسم IT - {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 30
    
    # العناوين
    headers = [
        'ID',
        'عنوان المهمة',
        'الوصف',
        'المشروع',
        'المكلف',
        'الحالة',
        'الأولوية',
        'المهارات المطلوبة',
        'تاريخ الإنشاء',
        'تاريخ الاستحقاق',
        'تاريخ التحديث'
    ]
    
    ws.append([])
    ws.append(headers)
    
    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    ws.row_dimensions[3].height = 25
    
    # البيانات
    tasks = Task.objects.all().select_related('project', 'assignee').order_by('-created_at')
    
    for task in tasks:
        # ✅ معالجة الحقول بشكل آمن
        task_id = task.id
        task_title = str(task.title) if task.title else 'لا يوجد'
        task_description = str(task.description) if task.description else 'لا يوجد'
        project_name = str(task.project.name) if task.project else '⭐ مهمة مستقلة'
        assignee_name = str(task.assignee.username) if task.assignee else 'غير محدد'
        
        # ✅ معالجة الحالة والأولوية
        try:
            status_display = task.get_status_display()
        except:
            status_display = str(task.status) if task.status else 'غير محدد'
        
        try:
            priority_display = task.get_priority_display()
        except:
            priority_display = str(task.priority) if task.priority else 'غير محدد'
        
        # ✅ معالجة المهارات المطلوبة
        required_skills = str(task.required_skills) if task.required_skills else 'لا يوجد'
        
        # ✅ معالجة التواريخ
        created_at = task.created_at.strftime('%Y-%m-%d %H:%M') if task.created_at else 'غير محدد'
        due_date = task.due_date.strftime('%Y-%m-%d') if task.due_date else 'غير محدد'
        updated_at = task.updated_at.strftime('%Y-%m-%d %H:%M') if task.updated_at else 'غير محدد'
        
        row_data = [
            task_id,
            task_title,
            task_description,
            project_name,
            assignee_name,
            status_display,
            priority_display,
            required_skills,
            created_at,
            due_date,
            updated_at
        ]
        
        ws.append(row_data)
        
        current_row = ws.max_row
        for cell in ws[current_row]:
            cell.alignment = data_alignment
            cell.border = border
            
            # تلوين حسب الحالة
            if cell.column == 6:  # عمود الحالة
                if task.status == 'completed':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif task.status == 'in_progress':
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
            
            # تلوين حسب الأولوية
            if cell.column == 7:  # عمود الأولوية
                if task.priority == 'high':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                elif task.priority == 'medium':
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # الإحصائيات
    ws.append([])
    ws.append([])
    
    stats_row = ws.max_row
    ws.merge_cells(f'A{stats_row}:D{stats_row}')
    ws[f'A{stats_row}'] = '📊 الإحصائيات'
    ws[f'A{stats_row}'].font = Font(size=12, bold=True, color="667eea")
    ws[f'A{stats_row}'].alignment = Alignment(horizontal='right')
    
    ws.append([
        'إجمالي المهام:',
        tasks.count(),
        '',
        'المهام المكتملة:',
        tasks.filter(status='completed').count()
    ])
    
    ws.append([
        'قيد التنفيذ:',
        tasks.filter(status='in_progress').count(),
        '',
        'لم تبدأ:',
        tasks.filter(status='to_do').count()
    ])
    
    # عرض الأعمدة
    column_widths = {
        'A': 8,
        'B': 25,
        'C': 35,
        'D': 20,
        'E': 15,
        'F': 12,
        'G': 12,
        'H': 25,
        'I': 18,
        'J': 15,
        'K': 18,
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'IT_tasks_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

@login_required
def export_submissions_excel(request):
    """تصدير التسليمات إلى Excel"""
    profile = Profile.objects.filter(user=request.user).first()
    if not profile or profile.role != "manager":
        messages.error(request, "غير مصرح لك.")
        return redirect('dashboard')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "التسليمات"
    
    # التنسيقات
    title_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # العنوان الرئيسي
    ws.merge_cells('A1:J1')
    ws['A1'] = f'📤 تقرير التسليمات - قسم IT - {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 30
    
    # العناوين
    headers = [
        'ID',
        'المهمة',
        'المشروع',
        'الموظف',
        'ملاحظات الموظف',
        'الحالة',
        'التقييم',
        'ملاحظات المدير',
        'تاريخ التسليم',
        'تاريخ المراجعة'
    ]
    
    ws.append([])
    ws.append(headers)
    
    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    ws.row_dimensions[3].height = 25
    
    # البيانات
    submissions = Submission.objects.all().select_related(
        'task',
        'task__project',
        'employee',
        'reviewed_by'
    ).order_by('-submitted_at')
    
    for sub in submissions:
        row_data = [
            sub.id,
            sub.task.title,
            sub.task.project.name if sub.task.project else '⭐ مهمة مستقلة',
            sub.employee.get_full_name() or sub.employee.username,
            sub.notes or 'لا يوجد',
            sub.get_status_display(),
            f"{sub.rating} ⭐" if sub.rating else 'لم يتم التقييم',
            sub.manager_comment or 'لا يوجد',
            sub.submitted_at.strftime('%Y-%m-%d %H:%M'),
            sub.reviewed_at.strftime('%Y-%m-%d %H:%M') if sub.reviewed_at else 'لم تتم المراجعة'
        ]
        
        ws.append(row_data)
        
        current_row = ws.max_row
        for cell in ws[current_row]:
            cell.alignment = data_alignment
            cell.border = border
            
            # تلوين حسب الحالة
            if cell.column == 6:
                if sub.status == 'approved':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif sub.status == 'rejected':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            # تلوين حسب التقييم
            if cell.column == 7 and sub.rating:
                if sub.rating >= 4:
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif sub.rating >= 3:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # الإحصائيات
    ws.append([])
    ws.append([])
    
    stats_row = ws.max_row
    ws.merge_cells(f'A{stats_row}:D{stats_row}')
    ws[f'A{stats_row}'] = '📊 الإحصائيات'
    ws[f'A{stats_row}'].font = Font(size=12, bold=True, color="667eea")
    ws[f'A{stats_row}'].alignment = Alignment(horizontal='right')
    
    ws.append([
        'إجمالي التسليمات:',
        submissions.count(),
        '',
        'الموافق عليها:',
        submissions.filter(status='approved').count()
    ])
    
    ws.append([
        'المرفوضة:',
        submissions.filter(status='rejected').count(),
        '',
        'قيد المراجعة:',
        submissions.filter(status='under_review').count()
    ])
    
    # متوسط التقييم
    approved_submissions = submissions.filter(status='approved', rating__isnull=False)
    if approved_submissions.exists():
        avg_rating = sum(s.rating for s in approved_submissions) / approved_submissions.count()
        ws.append([
            'متوسط التقييم:',
            f"{avg_rating:.2f} ⭐",
            '',
            '',
            ''
        ])
    
    # عرض الأعمدة
    column_widths = {
        'A': 8,
        'B': 25,
        'C': 20,
        'D': 18,
        'E': 30,
        'F': 15,
        'G': 12,
        'H': 30,
        'I': 18,
        'J': 18,
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'IT_submissions_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response